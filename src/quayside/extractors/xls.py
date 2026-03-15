"""Extract prices from XLS/XLSX files.

Handles two cases:
1. Known formats (matching existing scraper layouts) — deterministic parsing
2. Unknown formats — uses AI extraction as fallback
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from quayside.models import PriceRecord

logger = logging.getLogger(__name__)


def extract_prices(file_path: Path, port: str, date: str) -> list[PriceRecord]:
    """Extract price records from an XLS or XLSX file."""
    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        return _extract_xlsx(file_path, port, date)
    else:
        return _extract_xls(file_path, port, date)


def _extract_xlsx(file_path: Path, port: str, date: str) -> list[PriceRecord]:
    """Parse XLSX using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    records = []
    now = datetime.now().isoformat()

    # Try to detect the header row
    header_row = None
    headers = {}
    for row in ws.iter_rows(max_row=20, values_only=False):
        cells = [c.value for c in row]
        cell_lower = [str(c).lower().strip() if c else "" for c in cells]

        # Look for species/price columns
        if any(kw in " ".join(cell_lower) for kw in ("species", "price", "avg", "average")):
            header_row = row[0].row
            for i, val in enumerate(cell_lower):
                if "species" in val or "fish" in val:
                    headers["species"] = i
                elif "grade" in val:
                    headers["grade"] = i
                elif "low" in val or "min" in val:
                    headers["low"] = i
                elif "high" in val or "max" in val:
                    headers["high"] = i
                elif "avg" in val or "average" in val or "mean" in val or "price" in val and "avg" not in headers:
                    headers["avg"] = i
            break

    if not header_row or "species" not in headers:
        wb.close()
        logger.info("Could not detect headers in XLSX — falling back to AI extraction")
        from quayside.extractors.ai import extract_prices as ai_extract
        return ai_extract(file_path, port, date)

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        species_val = row[headers["species"]] if headers.get("species") is not None else None
        if not species_val or not str(species_val).strip():
            continue

        species = str(species_val).strip()
        grade = str(row[headers["grade"]]).strip() if headers.get("grade") is not None and row[headers["grade"]] else ""
        low = _to_float(row[headers["low"]]) if headers.get("low") is not None else None
        high = _to_float(row[headers["high"]]) if headers.get("high") is not None else None
        avg = _to_float(row[headers["avg"]]) if headers.get("avg") is not None else None

        # Skip rows with no price data at all
        if low is None and high is None and avg is None:
            continue

        # Calculate avg from low/high if missing
        if avg is None and low is not None and high is not None:
            avg = round((low + high) / 2, 2)

        records.append(PriceRecord(
            date=date, port=port, species=species, grade=grade,
            price_low=low, price_high=high, price_avg=avg, scraped_at=now,
        ))

    wb.close()
    logger.info("Extracted %d prices from XLSX for %s", len(records), port)
    return records


def _extract_xls(file_path: Path, port: str, date: str) -> list[PriceRecord]:
    """Parse XLS using xlrd."""
    import xlrd

    wb = xlrd.open_workbook(str(file_path))
    ws = wb.sheet_by_index(0)
    records = []
    now = datetime.now().isoformat()

    # Same header detection logic
    header_row = None
    headers = {}
    for row_idx in range(min(20, ws.nrows)):
        cells = [str(ws.cell_value(row_idx, c)).lower().strip() for c in range(ws.ncols)]
        if any(kw in " ".join(cells) for kw in ("species", "price", "avg", "average")):
            header_row = row_idx
            for i, val in enumerate(cells):
                if "species" in val or "fish" in val:
                    headers["species"] = i
                elif "grade" in val:
                    headers["grade"] = i
                elif "low" in val or "min" in val:
                    headers["low"] = i
                elif "high" in val or "max" in val:
                    headers["high"] = i
                elif "avg" in val or "average" in val or "mean" in val or "price" in val and "avg" not in headers:
                    headers["avg"] = i
            break

    if not header_row or "species" not in headers:
        logger.info("Could not detect headers in XLS — falling back to AI extraction")
        from quayside.extractors.ai import extract_prices as ai_extract
        return ai_extract(file_path, port, date)

    for row_idx in range(header_row + 1, ws.nrows):
        species_val = ws.cell_value(row_idx, headers["species"])
        if not species_val or not str(species_val).strip():
            continue

        species = str(species_val).strip()
        grade = ""
        if headers.get("grade") is not None:
            grade = str(ws.cell_value(row_idx, headers["grade"])).strip()

        low = _to_float(ws.cell_value(row_idx, headers["low"])) if headers.get("low") is not None else None
        high = _to_float(ws.cell_value(row_idx, headers["high"])) if headers.get("high") is not None else None
        avg = _to_float(ws.cell_value(row_idx, headers["avg"])) if headers.get("avg") is not None else None

        if low is None and high is None and avg is None:
            continue
        if avg is None and low is not None and high is not None:
            avg = round((low + high) / 2, 2)

        records.append(PriceRecord(
            date=date, port=port, species=species, grade=grade,
            price_low=low, price_high=high, price_avg=avg, scraped_at=now,
        ))

    logger.info("Extracted %d prices from XLS for %s", len(records), port)
    return records


def _to_float(val) -> float | None:
    """Safely convert a cell value to float."""
    if val is None or val == "":
        return None
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return None
