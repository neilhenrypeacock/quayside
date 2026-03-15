"""Generate port-specific XLSX upload templates.

Creates a pre-formatted Excel workbook that port clerks fill in each morning.
Species names come from a dropdown (canonical list), date auto-fills,
price columns are formatted for GBP. Much more usable than a web form.
"""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from quayside.species import get_all_canonical_names

logger = logging.getLogger(__name__)

OUTPUT_DIR = Path(__file__).resolve().parents[2] / "output" / "templates"

# Quayside colours
_TIDE = "1A3A4A"
_PAPER = "F5F0E8"
_WHITE = "FFFFFF"
_BORDER_COLOR = "E0DDD5"

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color=_WHITE)
_HEADER_FILL = PatternFill(start_color=_TIDE, end_color=_TIDE, fill_type="solid")
_BODY_FONT = Font(name="Calibri", size=11)
_DATE_FONT = Font(name="Calibri", size=11, bold=True, color=_TIDE)
_THIN_BORDER = Border(
    bottom=Side(style="thin", color=_BORDER_COLOR),
)
_ALT_FILL = PatternFill(start_color=_PAPER, end_color=_PAPER, fill_type="solid")


def generate_template(port_name: str, port_code: str, num_rows: int = 40) -> Path:
    """Generate a port-specific XLSX upload template.

    Args:
        port_name: Display name (e.g. "Grimsby")
        port_code: 3-letter code (e.g. "GRM")
        num_rows: Number of blank price rows to include

    Returns:
        Path to the generated XLSX file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prices"

    # --- Title row ---
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"{port_name} — Daily Prices"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=_TIDE)
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Date row ---
    ws["A2"] = "Date:"
    ws["A2"].font = _DATE_FONT
    ws["B2"] = date.today().isoformat()
    ws["B2"].font = _DATE_FONT
    ws["B2"].number_format = "YYYY-MM-DD"
    ws.row_dimensions[2].height = 22

    # --- Instructions ---
    ws.merge_cells("A3:F3")
    ws["A3"] = "Fill in your prices below, then email this file to prices@quayside.fish"
    ws["A3"].font = Font(name="Calibri", size=10, italic=True, color="888888")
    ws.row_dimensions[3].height = 20

    # --- Header row (row 4) ---
    headers = ["Species", "Grade", "Low (£/kg)", "High (£/kg)", "Avg (£/kg)", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center" if col_idx >= 3 else "left")
    ws.row_dimensions[4].height = 24

    # --- Column widths ---
    ws.column_dimensions["A"].width = 22  # Species
    ws.column_dimensions["B"].width = 10  # Grade
    ws.column_dimensions["C"].width = 14  # Low
    ws.column_dimensions["D"].width = 14  # High
    ws.column_dimensions["E"].width = 14  # Avg
    ws.column_dimensions["F"].width = 18  # Notes

    # --- Species dropdown validation ---
    species_list = get_all_canonical_names()
    # openpyxl formula validation — comma-separated list in quotes
    species_str = ",".join(f'"{s}"' for s in species_list)

    # If the list is too long for inline validation, use a hidden sheet
    if len(species_str) > 255:
        species_sheet = wb.create_sheet("_Species")
        for i, sp in enumerate(species_list, 1):
            species_sheet.cell(row=i, column=1, value=sp)
        species_sheet.sheet_state = "hidden"

        species_range = f"_Species!$A$1:$A${len(species_list)}"
        dv = DataValidation(
            type="list",
            formula1=species_range,
            allow_blank=True,
        )
    else:
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(species_list)}"',
            allow_blank=True,
        )

    dv.error = "Please select a species from the list, or type a new one"
    dv.errorTitle = "Species"
    dv.prompt = "Select a species or type your own"
    dv.promptTitle = "Species"
    dv.showErrorMessage = False  # Allow custom species not in the list

    # --- Data rows ---
    first_data_row = 5
    for row_idx in range(first_data_row, first_data_row + num_rows):
        # Alternating row fill
        if (row_idx - first_data_row) % 2 == 1:
            for col_idx in range(1, 7):
                ws.cell(row=row_idx, column=col_idx).fill = _ALT_FILL

        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER

        # Price columns: number format
        for col_idx in [3, 4, 5]:
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="right")

    # Apply species validation to column A
    dv_range = f"A{first_data_row}:A{first_data_row + num_rows - 1}"
    dv.add(dv_range)
    ws.add_data_validation(dv)

    # --- Freeze panes (header always visible) ---
    ws.freeze_panes = "A5"

    # --- Print setup ---
    ws.print_title_rows = "4:4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    # Save
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"quayside_template_{port_code.lower()}.xlsx"
    path = OUTPUT_DIR / filename
    wb.save(path)

    logger.info("Generated template: %s (%d species in dropdown)", path, len(species_list))
    return path
