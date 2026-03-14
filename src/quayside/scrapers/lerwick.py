"""Scrape Lerwick/Shetland fish prices and landings from shetlandauction.com.

The SSA portal generates daily price data as downloadable XLSX files:
    https://ssawebportal.azurewebsites.net/daily-prices-generate.php?thisdate=YYYY-MM-DD

XLSX format (sheet "Daily Prices - YYYY-MM-DD"):
    Row 1:  "FISH MARKET PRICES SHETLAND"
    Row 2:  "WEEKLY TOTAL - N BOXES"
    Row 4:  "Friday 13th March 2026"
    Row 6:  "9 BOATS LANDED 1877 BOXES"
    Row 13: Headers: SPECIES | GRADE | VOLUME (kgs) | MAXIMUM PRICE (£/kg) | AVERAGE PRICE (£/kg)
    Row 17+: Data rows (with blank rows between species groups)

Species use abbreviated names (COD, HADD, MONK, etc.) which are mapped
to full names for consistency with other ports.
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from io import BytesIO

import requests

from quayside.models import PriceRecord

logger = logging.getLogger(__name__)

PORT = "Lerwick"
PRICES_URL = "https://ssawebportal.azurewebsites.net/daily-prices-generate.php"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0 Safari/537.36",
}

# Map SSA abbreviated species names to full names
_SPECIES_MAP = {
    "BLL": "Brill",
    "CATS": "Catfish",
    "COD": "Cod",
    "DOGS": "Dogfish",
    "GUR": "Gurnard",
    "HADD": "Haddock",
    "HAKE": "Hake",
    "HAL": "Halibut",
    "HR": "Haddock Round",
    "JD": "John Dory",
    "LEM": "Lemons",
    "LING": "Ling",
    "LYTH": "Lythe/Pollack",
    "MEG": "Megrim",
    "MEG BR": "Megrim Bruised",
    "MIX": "Mixed",
    "MONK": "Monks",
    "PLE": "Plaice",
    "PRNS": "Prawns",
    "ROES": "Roes",
    "SAI": "Saithe",
    "SK": "Skate",
    "SKM": "Skate Medium",
    "SKR": "Skate Round",
    "SKS": "Skate Small",
    "SQU": "Squid",
    "TUR": "Turbot",
    "WHIT": "Whiting",
    "WR": "Whiting Round",
    "WTS": "Witches",
}


def scrape_prices(
    target_date: str | None = None,
    xlsx_bytes: bytes | None = None,
) -> list[PriceRecord]:
    """Scrape Lerwick prices from the SSA portal XLSX download."""
    if target_date is None:
        target_date = datetime.now().strftime("%Y-%m-%d")

    if xlsx_bytes is None:
        url = f"{PRICES_URL}?thisdate={target_date}"
        logger.info("Fetching Lerwick prices: %s", url)
        try:
            resp = requests.get(url, headers=HEADERS, timeout=30)
            resp.raise_for_status()
            xlsx_bytes = resp.content
        except requests.RequestException as e:
            logger.warning("Could not fetch Lerwick prices: %s", e)
            return []

        # Check we got an XLSX (starts with PK zip header), not an error page
        if not xlsx_bytes.startswith(b"PK"):
            logger.warning("Lerwick response is not an XLSX file")
            return []

    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl required for Lerwick XLSX parsing — pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Extract date from the sheet (row 4 typically: "Friday 13th March 2026")
    date = _extract_date(ws) or target_date

    scraped_at = datetime.utcnow().isoformat()
    records = []

    for row in ws.iter_rows(min_row=1, values_only=True):
        species_raw = row[0] if row[0] else None
        grade_raw = row[1] if len(row) > 1 else None
        max_price_raw = row[3] if len(row) > 3 else None
        avg_price_raw = row[4] if len(row) > 4 else None

        # Skip non-data rows
        if not species_raw or species_raw in ("SPECIES", ""):
            continue
        if str(species_raw).startswith(("FISH", "WEEKLY", "DAILY")):
            continue

        # Parse prices
        max_price = _parse_num(max_price_raw)
        avg_price = _parse_num(avg_price_raw)

        if max_price is None and avg_price is None:
            continue

        species = _SPECIES_MAP.get(str(species_raw).strip(), str(species_raw).strip().title())
        grade = str(grade_raw).strip() if grade_raw is not None else "ALL"

        records.append(
            PriceRecord(
                date=date,
                port=PORT,
                species=species,
                grade=grade,
                price_low=None,
                price_high=max_price,
                price_avg=avg_price,
                scraped_at=scraped_at,
            )
        )

    wb.close()
    logger.info("Scraped %d price records for %s on %s", len(records), PORT, date)
    return records


def _extract_date(ws) -> str | None:
    """Extract date from the first few rows of the worksheet."""
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        text = str(row[0]) if row[0] else ""
        m = re.search(r"(\d+)(?:st|nd|rd|th)\s+(\w+)\s+(\d{4})", text)
        if m:
            try:
                dt = datetime.strptime(
                    f"{m.group(1)} {m.group(2)} {m.group(3)}", "%d %B %Y"
                )
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def _parse_num(val) -> float | None:
    """Parse a numeric value, returning None for empty/invalid."""
    if val is None:
        return None
    try:
        v = float(val)
        return round(v, 2) if v > 0 else None
    except (ValueError, TypeError):
        return None
